# -*- utf-8 -*-
#@Time    :2019/3/916:04
#@Author  :无邪
#@File    :read_excel.py
#@Software:PyCharm
from Item.common.read_conf import Readconf
from openpyxl import load_workbook
from Item.common.path_jion import SplitPath


class Readexcel:
    """读取excel用例的类"""
    def __init__(self,filename):
        self.filename=filename#将文件名赋值给对象
        self.wb=load_workbook(filename=self.filename)#打开工作薄
        self.sheet=self.wb["Sheet1"]#定位表单
    def get_data(self):
        self.sheet2=self.wb["Sheet2"]
        mobile=self.sheet2.cell(1,2).value
        return mobile
    def update_tel(self,value):
        "更新"
        self.sheet2=self.wb["Sheet2"]
        self.sheet2.cell(1,2).value=value
        self.wb.save(self.filename)
    def readcase(self):
        """读取用例"""
        #获取配置文件中用例执行控制开关button
        button=Readconf(SplitPath().join_confname()).read_button()
        if button=="1":
            mobile=self.get_data()#获取初始化手机号
            single_one=[]
            for i in range(1,self.sheet.max_row):#读取全部用例"
                single_case={}
                single_case["Case_id"] = self.sheet.cell(i+1,1).value
                single_case["module"]=self.sheet.cell(i+1,2).value
                single_case["url"]=self.sheet.cell(i+1,5).value
                single_case["Params"] = self.sheet.cell(i + 1, 6).value
                single_case["sql"] = self.sheet.cell(i + 1, 7).value
                single_case["ExpectedResult"] = self.sheet.cell(i + 1, 9).value
                if self.sheet.cell(i+1,6).value.find("tel")!=-1:
                    #如果手机号存在，就替换
                    newvalue=self.sheet.cell(i+1,6).value.replace('tel',str(mobile))
                #-------------------------------------替换参数方法1---------------------------------------------------------
                # elif self.sheet.cell(i + 1, 6).value.find("%re_call") != -1:
                #         # 将手机号参数化，从getdata模块中获取
                #     newvalue = self.sheet.cell(i + 1, 6).value.replace("%re_call",str(ZhengZ().use_tel()))
                # elif self.sheet.cell(i + 1, 6).value.find("%re_memberId") != -1:
                #         # 将正常用例的用户id参数化，从getdata模块中获取
                #     newvalue = self.sheet.cell(i + 1, 6).value.replace("%re_memberId", str(str(ZhengZ().use_id())))
                #----------------------------------------替换参数方法2----------------------------------------------------
                # elif self.sheet.cell(i + 1, 6).value.find("call") != -1:
                #         # 将手机号参数化，从getdata模块中获取
                #     newvalue = self.sheet.cell(i + 1, 6).value.replace("call", str(getattr(GetData, "call")))
                # elif self.sheet.cell(i + 1, 6).value.find("userid") != -1:
                #         # 将正常用例的用户id参数化，从getdata模块中获取
                #     newvalue = self.sheet.cell(i + 1, 6).value.replace("userid", str(getattr(GetData, "memberId")))
                #---------------------------------------------------------------------------------------------------------
                else:
                        # 不替换
                    newvalue = self.sheet.cell(i + 1, 6).value
                single_case["Params"] = newvalue
                single_one.append(single_case)
            self.update_tel(mobile+1)
            return single_one
        else:
            mobile=self.get_data()
            single_one = []
            for i in eval(button):#选择单个用例
                single_case = {}
                single_case["Case_id"] = self.sheet.cell(i+1,1).value
                single_case["module"] = self.sheet.cell(i + 1, 2).value
                single_case["url"] = self.sheet.cell(i+1,5).value
                single_case["Params"] = self.sheet.cell(i+1,6).value
                single_case["sql"] = self.sheet.cell(i + 1, 7).value
                single_case["ExpectedResult"] = self.sheet.cell(i+1,9).value
                if self.sheet.cell(i+1, 6).value.find("tel") !=-1:
                    # 如果手机号存在，就替换
                    newvalue =self.sheet.cell(i+1, 6).value.replace("tel",str(mobile))
                # -------------------------------------替换参数方法1---------------------------------------------------------
                # elif self.sheet.cell(i + 1, 6).value.find("%re_call") != -1:
                #     # 将手机号参数化，从getdata模块中获取
                #     newvalue = self.sheet.cell(i + 1, 6).value.replace(
                #         "%re_call", str(ZhengZ().use_tel()))
                # elif self.sheet.cell(i + 1, 6).value.find("%re_memberId") != -1:
                #     # 将正常用例的用户id参数化，从getdata模块中获取
                #     newvalue = self.sheet.cell(i + 1, 6).value.replace(
                #         "%re_memberId", str(str(ZhengZ().use_id())))
                # -------------------------------------替换参数方法2---------------------------------------------------------
                # elif self.sheet.cell(i + 1, 6).value.find("call")!=-1:
                #     #将手机号参数化，从getdata模块中获取
                #     newvalue = self.sheet.cell(i + 1, 6).value.replace("call",str(getattr(GetData,"call")))
                # elif self.sheet.cell(i + 1, 6).value.find("userid")!=-1:
                #     #将正常用例的用户id参数化，从getdata模块中获取
                #     newvalue = self.sheet.cell(i + 1, 6).value.replace("userid",str(getattr(GetData,"memberId")))
                #----------------------------------------------------------------------------------------------------
                else:
                    # 不替换
                    newvalue = self.sheet.cell(i+1, 6).value
                single_case["Params"]=newvalue
                single_one.append(single_case)
                self.update_tel(mobile+ 1)
            return single_one
    def writeexcel(self,row,column,value):
        self.sheet.cell(row,column).value=value
        self.wb.save(self.filename)
if __name__ == '__main__':
    readexcel=Readexcel(SplitPath().join_excelname())
    print(readexcel.readcase())